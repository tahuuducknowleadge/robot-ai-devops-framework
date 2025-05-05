from openpyxl import load_workbook
from robot.api.deco import keyword
from robot.libraries.BuiltIn import BuiltIn
from faker import Faker

# Khởi tạo Faker (mặc định là en_US, có thể đổi thành vi_VN nếu cần)
fake = Faker('en_US')  # Hoặc fake = Faker('vi_VN') cho dữ liệu tiếng Việt

# Dictionary ánh xạ lệnh faker với hàm tương ứng
FAKER_MAPPINGS = {
    "{faker_name}": fake.name,
    "{faker_email}": fake.email,
    "{faker_phone}": fake.phone_number,
    "{faker_address}": fake.address,
    "{faker_text}": lambda: fake.text(max_nb_chars=50),  # Văn bản ngắn
    "{faker_number}": lambda: fake.random_int(min=1, max=1000),  # Số ngẫu nhiên
    # Thêm các loại dữ liệu khác nếu cần
}

# Các action không cần value (chạy nếu value == "TRUE")
NO_VALUE_ACTIONS = {"clickButton", "checkCheckbox"}

def normalize_locator(raw_locator: str) -> str:
    """Tự động thêm prefix 'xpath=' nếu cần."""
    raw_locator = raw_locator.strip()
    if "=" in raw_locator.split()[0]:
        return raw_locator
    if raw_locator.startswith(("//", ".//", "(")):
        return f"xpath={raw_locator}"
    return raw_locator

def process_faker_value(value_str: str) -> str:
    """Xử lý giá trị {faker_...} để sinh dữ liệu động."""
    value_str = value_str.strip()
    if value_str in FAKER_MAPPINGS:
        return str(FAKER_MAPPINGS[value_str]())  # Gọi hàm Faker tương ứng
    return value_str  # Giữ nguyên nếu không phải lệnh Faker

@keyword("Run Horizontal Test From Excel")
def run_horizontal_test_from_excel(file_path, test_case_name, sheet_name="Sheet1"):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    try:
        tc_index = header.index(test_case_name)
    except ValueError:
        raise Exception(f"❌ Không tìm thấy test case '{test_case_name}' trong file Excel.")

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        action = row[0]
        locator = row[1]
        value = row[tc_index] if tc_index < len(row) else ""

        if not action:
            continue

        action = action.strip()
        locator = str(locator).strip() if locator else ""
        value_str = str(value).strip() if value is not None else ""
        normalized_value = value_str.upper()

        is_no_value_action = action in NO_VALUE_ACTIONS

        # 🟡 Nếu là action không cần value → chỉ chạy khi value == TRUE
        if is_no_value_action:
            if normalized_value != "TRUE":
                continue
            if locator:
                locator = normalize_locator(locator)
                locator = BuiltIn().replace_variables(locator)
                BuiltIn().log(f"▶️ [Row {row_index}] Action: {action} | Locator: {locator}", level="INFO")
                try:
                    BuiltIn().run_keyword(action, locator)
                except Exception as e:
                    BuiltIn().log(f"❌ [Row {row_index}] Failed: {action} with locator {locator}: {str(e)}", level="ERROR")
                    raise
            continue

        # 🟢 Với các action có value (action cần giá trị)
        if not locator and not value_str:
            BuiltIn().log(f"⏩ [Row {row_index}] Skipped: No locator and no value for action {action}", level="INFO")
            continue

        args = []
        if locator:
            locator = normalize_locator(locator)
            locator = BuiltIn().replace_variables(locator)
            args.append(locator)

        if action == "enterText":  # Đảm bảo enterText luôn có 2 tham số
            value_str = value_str if value_str else ""  # Dùng chuỗi rỗng làm mặc định
            # Xử lý giá trị {faker_...} trước khi thay biến
            value_str = process_faker_value(value_str)
            value_str = BuiltIn().replace_variables(value_str)
            args.append(value_str)
        elif value_str:  # Với các action khác, chỉ thêm value nếu không trống
            # Xử lý giá trị {faker_...} trước khi thay biến
            value_str = process_faker_value(value_str)
            value_str = BuiltIn().replace_variables(value_str)
            args.append(value_str)

        BuiltIn().log(f"▶️ [Row {row_index}] Action: {action} | Args: {args}", level="INFO")
        try:
            BuiltIn().run_keyword(action, *args)
        except Exception as e:
            BuiltIn().log(f"❌ [Row {row_index}] Failed: {action} with args {args}: {str(e)}", level="ERROR")
            raise